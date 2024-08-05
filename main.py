import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import re
import os
import platform
import subprocess
from PIL import Image, ImageTk  # Necesario para usar imágenes en formatos diferentes a .ico

# Comprobación de si existe el archivo datos.xlsx
nombre_archivo = 'datos.xlsx'

if os.path.exists(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
    # Crear el Excel
    wb = Workbook()
    ws = wb.active

    # Creamos la primera fila y aplicamos negrita
    header = ["Nombre", "Edad", "Email", "Teléfono", "Dirección", "Transacción"]
    ws.append(header)
    for cell in ws[1]:  # Iterar sobre la primera fila
        cell.font = Font(bold=True)

# Función para limpiar los campos del formulario
def limpiar_formulario():
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)
    entry_transaccion.delete(0, tk.END)

# Función para sumar la columna de transacciones y guardar el total en la columna 8
def sumar_transacciones():
    total = 0
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
        if row[0] is not None:
            total += row[0]

    ws['H1'] = "Total Transacciones"
    ws['H1'].font = Font(bold=True)
    ws['H2'] = total

# Función para el botón guardar datos
def guardar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefono = entry_telefono.get()
    direccion = entry_direccion.get()
    transaccion = entry_transaccion.get()
    
    # Validamos las entrada para que ninguna esté vacía
    if not nombre or not edad or not email or not telefono or not direccion or not transaccion:
        messagebox.showwarning(title="Advertencia", message="Todos los campos son obligatorios")
        return
    
    # Comprobamos que la edad y el teléfono son números
    try:
        edad = int(edad)
        telefono = int(telefono)
        transaccion = int(transaccion)
    except ValueError:
        messagebox.showwarning(title="Advertencia", message="Edad, Teléfono y Transacción deben ser números")
        return
    
    # Validamos el formato del email (es una comprobación básica "texto@texto.texto")
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showwarning(title="Advertencia", message="El correo electrónico no es válido")
        return
    
    ws.append([nombre, edad, email, telefono, direccion, transaccion])

    # Sumamos las transacciones y guardamos el total
    sumar_transacciones()

    # Guardamos el archivo Excel
    wb.save(nombre_archivo)
    messagebox.showinfo(title="Información", message="Datos guardados con éxito")
    
    # Limpiamos los campos del formulario
    limpiar_formulario()

# Función para abrir el archivo Excel
def abrir_archivo():
    # En caso de que el archivo Excel no exista todavía
    if not os.path.exists(nombre_archivo):
        messagebox.showwarning(title="Advertencia", message="El archivo Excel aún no ha sido creado.")
        return
    
    if platform.system() == "Windows":
        os.startfile(nombre_archivo)
    elif platform.system() == "Darwin":  # macOS
        subprocess.call(["open", nombre_archivo])
    else:  # Linux and other Unix-like systems
        subprocess.call(["xdg-open", nombre_archivo])

root = tk.Tk()
root.title("Formulario de Entrada de Datos")  # Cambia este texto para modificar el tooltip

# Hacer la ventana no redimensionable
root.resizable(False, False)

# Configurar el ícono
try:
    # Para Windows y macOS, usa iconbitmap para .ico o iconphoto para otros formatos
    icon = Image.open("icono.png")  # Usa PIL para abrir el ícono
    root.iconphoto(True, ImageTk.PhotoImage(icon))
except Exception as e:
    print(f"No se pudo cargar el ícono: {e}")

# Estilos del formulario de entrada de datos
root.configure(bg='#1e8a9c')
label_style = {"bg": "#1e8a9c", "fg": "white"}
entry_style = {"bg": "#D3D3D3", "fg": "black"}

# Etiquetas y posiciones
label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

label_edad = tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)

label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

label_telefono = tk.Label(root, text="Teléfono", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

label_direccion = tk.Label(root, text="Dirección", **label_style)
label_direccion.grid(row=4, column=0, padx=10, pady=5)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)

label_transaccion = tk.Label(root, text="Transacción", **label_style)
label_transaccion.grid(row=5, column=0, padx=10, pady=5)
entry_transaccion = tk.Entry(root, **entry_style)
entry_transaccion.grid(row=5, column=1, padx=10, pady=5)

# Botón guardar
boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos, bg='#196673', fg='white')
boton_guardar.grid(row=6, column=0, padx=10, pady=10)

# Botón abrir archivo
boton_abrir = tk.Button(root, text="Abrir Archivo", command=abrir_archivo, bg='#196673', fg='white')
boton_abrir.grid(row=6, column=1, padx=10, pady=10)

# Evitamos que se cierre la ventana
root.mainloop()